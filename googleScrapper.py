from openpyxl import load_workbook
import xlsxwriter
import os
import requests
import time
import re
from bs4 import BeautifulSoup
from selenium import webdriver



outfile = 'movieDataMore.xlsx'
try:
	os.remove(outfile)
except:
	print("newfile")

workbook = xlsxwriter.Workbook(outfile)
worksheet = workbook.add_worksheet()


wb = load_workbook(filename='movieData.xlsx', read_only=True)
ws = wb['Sheet1']
driver = webdriver.Chrome()
count = 1
for row in ws.rows:
	if count % 50 is 0:
		time.sleep(90)
	try:
		print(row[0].value)



		url  = "http://www.google.com/search?q=" + row[0].value + " movie"
		driver.get(url)
		html = driver.page_source
		soup = BeautifulSoup(html,"html.parser")

		genreRating= soup.find_all("div", class_="_gdf kno-fb-ctx")[0].text
		worksheet.write(count, 0, row[0].value)

		MPARating = genreRating.split(' ')[0]
		if re.search('[a-zA-Z]', MPARating):
			worksheet.write(count, 1, MPARating)
			#print(MPARating)
			genres = genreRating.split('‧')[1][1:]
			worksheet.write(count, 2, genres)
			#print(genres)
			duration = genreRating.split('‧')[2][1:]
			worksheet.write(count, 3, duration)
		else:
			#print(MPARating)
			genres = genreRating.split('‧')[0][1:]
			worksheet.write(count, 2, genres)
			#print(genres)
			duration = genreRating.split('‧')[1][1:]
			worksheet.write(count, 3, duration)

		#print(duration)
		try:
			ytTrailer= soup.find_all("a", class_="_glf ellip kno-fb-ctx")[0]['href']
			worksheet.write(count, 4, ytTrailer)
			#print (ytTrailer)
		except:
			print("sdfas")
		#ratings not always in the same space look for %
		ratings = soup.find_all("span", class_="_tvg")
		for rating in ratings:
			if rating.text.find("%") is not -1:
				worksheet.write(count, 5, rating.text)
				#print(rating.text)
				break;
		
		description = soup.find_all("div", class_="_cgc kno-fb-ctx")[0].text
		worksheet.write(count, 6, description)

		#print(description)#strip .... MORE
		try:
			releaseDate = soup.find_all(attrs = {'data-attrid':"kc:/film/film:theatrical region aware release date"})[0].text
			worksheet.write(count, 7, releaseDate)
			#print(releaseDate)#strip string
		except:
			try:
				releaseDate = soup.find_all(attrs = {'data-attrid':"kc:/film/film:initial theatrical regional release date"})[0].text
				worksheet.write(count, 7, releaseDate)
				#print(releaseDate)#strip string
			except:
				print("sdfsfd")

		director = soup.find_all(attrs = {'data-attrid':"kc:/film/film:director"})[0].text
		worksheet.write(count, 8, director)

		#print(director)#strip string
		try:
			screenplay = soup.find_all(attrs = {'data-attrid':"kc:/film/film:screenplay"})[0].text
			worksheet.write(count, 9, screenplay)

			#print(screenplay)#strip string
		except:
			print("sdfas")
		try:
			budget = soup.find_all(attrs = {'data-attrid':"hw:/collection/films:budget"})[0].text
			worksheet.write(count, 10, budget)

			#print(budget)#strip string
		except:
			print("sdfas")
		try:
			boxOffice = soup.find_all(attrs = {'data-attrid':"hw:/collection/films:box office"})[0].text
			worksheet.write(count, 11, boxOffice)

			#print(boxOffice)#strip string
		except:
			print("sdfas")

		similarMovies = soup.find_all("div", class_="_c4 _Dnh")[1].find_all("div", class_="fl ellip _NRl")
		movieStr = ""
		for movie in similarMovies:
			#print(movie.text + "|")
			movieStr += movie.text +"|"
		worksheet.write(count, 12, movieStr)
		try:
			actors = soup.find_all("div", class_="_c4 _Dnh")[3].find_all("div", class_="fl ellip _NRl")
			actorStr = ""
			for actor in actors:
				#print(actor.text + "|")
				actorStr += actor.text + "|"
			worksheet.write(count, 13, actorStr)
		except:
			print("sdfsfd")
	except:
		print(str(count))
	count+=1